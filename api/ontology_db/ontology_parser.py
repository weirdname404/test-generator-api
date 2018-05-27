# coding=utf-8

from api.ontology_db.entities.alloying_element import AlloyingElement
from api.ontology_db.db_config.base import Base, Session, engine
from api.ontology_db.entities.deoxidizing import DeoxidizingType
from api.ontology_db.entities.entity_class import EntityClass
from api.ontology_db.entities.gost import Gost
from api.ontology_db.entities.guid import Guid, ScaleGuid, ClassGuid
from api.ontology_db.entities.element_value import MinCarbonValue, MaxCarbonValue, MaxMarganeseValue, MinMarganeseValue
from api.ontology_db.entities.quality import QualityType
from api.ontology_db.entities.scales import Scale, ScaleValue
from api.ontology_db.entities.steel import Steel
from openpyxl import load_workbook

# generate database schema
Base.metadata.create_all(engine)

# start session
session = Session()

# ontology file
FILE_NAME = './ontology.xlsm'
ROW_MAX = 65

MAX_OBJS = 45
SUPPORTED_SCALES = 8
# obj_name and obj_class included
MAX_SCALES = SUPPORTED_SCALES + 2


# Parsing and inserting data about Scales from the ontology file ('Scales')
def parse_insert_scales(file_name):
    # Load in the ontology
    ws = load_workbook(file_name)['Scales']
    scale_object = ''

    for row in range(1, ROW_MAX):
        col = 1

        if ws.cell(row, col).value == 'Title':
            scale_name = ws.cell(row, col + 1).value
            scale_guid = ws.cell(row + 6, col + 1).value
            scale_object = Scale(scale_name, ScaleGuid(scale_guid))
            continue

        elif ws.cell(row, col).value == 'Values':
            while ws.cell(row, col + 1).value is not None:
                col += 1

                # Create Scale_value object and insert it in DB
                scale_value = ws.cell(row, col).value
                session.add(ScaleValue(scale_value, scale_object))

        else:
            continue

    # Save insert actions
    session.commit()
    session.close()

    print('\nScales are successfully parsed and moved to db')


def print_scales():
    ontology_scales = {}

    all_scales = session.query(Scale).all()

    print('\n### All scales:\n')

    for scale in all_scales:
        print(scale.name)
        print(scale.guid.name)
        print([value.value for value in scale.values], '\n')


class SavedObjects:
    dict = {}
    file_name = ''

    def __init__(self, file_name):
        self.file_name = file_name


# Parsing and inserting data about Steel Objects from the ontology file ('MVContext')
def parse_insert_objects(file_name):
    # Load in the ontology
    ws = load_workbook(file_name)['MVContext']

    # The main idea is to reuse already created objects, which will be contained in the dict 
    saved_objects = SavedObjects(file_name)

    for row in ws.iter_rows(min_row=2, max_col=MAX_SCALES + 1, max_row=MAX_OBJS):
        """
        The data is in the following order:

        0     1     2
        Name, GUID, Class,
        3     4                5
        Gost, DeoxidizingType, Quality,
        6                 7          8
        AlloyingElements, MinCarbon, MaxCarbon
        """

        # All data from 1 row
        row_data = [cell.value for cell in row]

        # Guid has one-to-one relation with the steel or unique, so there will be always N guid objects.
        steel_guid = Guid(row_data.pop(0))
        steel_name = row_data.pop(0)

        """
        The main idea of this part is to REUSE value objects instead of creating hundreds of new objects .

        At first we check if the object for current value was created in the past or not,
        Then we either create a new object for a current value and instantly use it or just use existing one,

        After objects were identified or created, we store them in the row_data_objs dict.
        Due to the fact that parser parses one row after another, 
        these created or identified objects will be used to create a Steel object for a current row.
        
        Ontology is parsed ONCE.
        """
        row_data_objs = define_values(saved_objects, row_data)

        # saving steel object in DB
        session.add(update_steel_object(Steel(steel_name, steel_guid), row_data_objs))

    # Save insert actions
    session.commit()
    session.close()

    print('Objects are successfully parsed and moved to db\n')


def define_values(saved_objects, row_data):
    value_objects = saved_objects.dict
    file_name = saved_objects.file_name
    row_data_objs = []

    for i in range(len(row_data)):
        current_value = row_data[i]

        # check if there are any keys, if not - create
        if i not in value_objects:
            value_objects[i] = {}

        # current_value is a "Element1, Element3, Element3 ..." - special treatment is required
        if i == 4:
            # Alloying_elements must be a list of objects - [obj1, obj2, obj3]
            row_data_objs.append(parse_alloying_elements(value_objects, current_value, i))
            continue

        # the case when value class was already instantiated in the past, so we should use existing object
        if current_value in value_objects[i]:
            row_data_objs.append(value_objects[i][current_value])

        # in other cases, let's create a value obj and use it in the future
        else:
            value_obj = define_object(i, file_name, current_value)
            value_objects[i][current_value] = value_obj
            row_data_objs.append(value_obj)

    return row_data_objs


# steel attributes init
def define_object(i, file_name, current_value):
    if i == 0:
        thesaurus = load_workbook(file_name)['Thesaurus']
        row, col = 2, 2

        while thesaurus.cell(row, col).value != current_value:
            row += 1

        class_guid = thesaurus.cell(row, col - 1).value
        value_obj = EntityClass(current_value, ClassGuid(class_guid))

    elif i == 1:
        value_obj = Gost(current_value)

    elif i == 2:
        value_obj = DeoxidizingType(current_value)

    elif i == 3:
        value_obj = QualityType(current_value)

    else:
        objects = {
            5: MinCarbonValue(current_value),
            6: MaxCarbonValue(current_value),
            7: MinMarganeseValue(current_value),
            8: MaxMarganeseValue(current_value)
        }

        value_obj = objects[i]

    return value_obj


# Object init and data updating
def update_steel_object(steel_object, row_data_objs):
    steel_object.entity_class = row_data_objs[0]
    steel_object.gost = row_data_objs[1]
    steel_object.deoxidizing_type = row_data_objs[2]
    steel_object.quality = row_data_objs[3]
    steel_object.alloying_elements = row_data_objs[4]
    steel_object.min_carbon_value = row_data_objs[5]
    steel_object.max_carbon_value = row_data_objs[6]
    steel_object.min_marganese_value = row_data_objs[7]
    steel_object.max_marganese_value = row_data_objs[8]

    return steel_object


def parse_alloying_elements(obj_dict, current_value, elements_key):
    element_objs_list = []
    alloying_elements = [i for i in current_value.split(', ')]

    # check, create, use
    for element in alloying_elements:
        # check if the key for a current element exists in a nested dict or not
        if element not in obj_dict[elements_key]:
            element_obj = AlloyingElement(element)
            obj_dict[elements_key][element] = element_obj

        element_objs_list.append(obj_dict[elements_key][element])

    return element_objs_list


def print_objects():
    all_objects = session.query(Steel) \
        .all()

    print('\n### All objects:\n')
    for steel in all_objects:
        print(steel.name)
        print(steel.guid.name)
        print(steel.gost.name)
        print(steel.entity_class.name, '(%s)' % (steel.entity_class.guid.name))
        print(steel.deoxidizing_type.name)
        print(steel.quality.name)
        print([element.name for element in steel.alloying_elements])
        print(steel.min_carbon_value.value)
        print(steel.max_carbon_value.value, '\n')


def clear_db_data(session):
    meta = Base.metadata
    print('\n')
    for table in reversed(meta.sorted_tables):
        print('Clear table %s' % table)
        session.execute(table.delete())
        session.commit()


def drop_tables(session):
    engine.execute("DROP TABLE steels CASCADE")
    engine.execute("DROP TABLE steels_alloying_elements CASCADE")
    engine.execute("DROP TABLE quality_types CASCADE")
    engine.execute("DROP TABLE min_carbon_values CASCADE")
    engine.execute("DROP TABLE max_carbon_values CASCADE")
    engine.execute("DROP TABLE guids CASCADE")
    engine.execute("DROP TABLE gosts CASCADE")
    engine.execute("DROP TABLE entity_classes CASCADE")
    engine.execute("DROP TABLE deoxidizing_types CASCADE")
    engine.execute("DROP TABLE alloying_elements CASCADE")
    engine.execute("DROP TABLE scale_guids CASCADE")
    engine.execute("DROP TABLE class_guids CASCADE")
    print('\nALL TABLES DROPPED\n')


def parse_ontology():
    clear_db_data(session)
    parse_insert_scales(FILE_NAME)
    parse_insert_objects(FILE_NAME)


def drop_ontology():
    drop_tables(session)
