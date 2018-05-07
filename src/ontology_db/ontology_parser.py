#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from .entities.db_config.base import Base, Session, engine
from .entities.scales import Scale, Scale_value
from .entities.steel import Steel
from .entities.alloying_element import Alloying_element
from .entities.deoxidizing import Deoxidizing_type
from .entities.entity_class import Entity_class
from .entities.gost import Gost
from .entities.guid import Guid
from .entities.max_carbon_value import Max_carbon_value
from .entities.min_carbon_value import Min_carbon_value
from .entities.quality import Quality_type

# generate database schema
Base.metadata.create_all(engine)

# start session
session = Session()

# ontology file
FILE_NAME = './ontology.xlsm'
ROW_MAX = 65


# parsing and inserting data about Scales from the ontology file
def parse_insert_scales(file_name):
    # Load in the workbook
    ws = load_workbook(file_name)['Scales']
    scale_name = ''

    for row in range(1, ROW_MAX):
        col = 1

        if ws.cell(row, col).value == 'Title':
            scale_name = ws.cell(row, col + 1).value
            continue

        elif ws.cell(row, col).value == 'Values':
            while ws.cell(row, col + 1).value is not None:
                col += 1

                # Create Scale_value object and insert it in DB
                scale_value = ws.cell(row, col).value
                session.add(Scale_value(scale_value, Scale(scale_name)))

        else:
            continue

    # Save insert actions
    session.commit()
    session.close()

    print('\nScales are successfully parsed and moved to db\n')
    test_scales()


def insert_objects(data):
    pass


# local tests with queries
def test_scales():
    ontology_scales = {}

    all_scales = session.query(Scale) \
        .all()

    # set of all scale names
    all_scale_names_set = set([i.name for i in all_scales])

    for scale_name in all_scale_names_set:
        scale_values = session.query(Scale_value) \
            .join(Scale) \
            .filter(Scale.name == scale_name) \
            .all()

        values = []

        for scale_value in scale_values:
            values.append(scale_value.value)

        ontology_scales[scale_name] = values

    print('\n### All scales:\n', ontology_scales)
    print('\n###Scale test finished\n')


def test_objects():
    pass


def clear_db_data(session):
    meta = Base.metadata
    for table in reversed(meta.sorted_tables):
        print('Clear table %s' % table)
        session.execute(table.delete())
        session.commit()


def ontology_parser():
    clear_db_data(session)
    parse_insert_scales(FILE_NAME)

    # insert_objects(parse_objects(FILE_NAME))

    # for row in ws.iter_rows(min_row=1, max_col=100, max_row=100):
    #     for cell in row:
    #         if cell.value == 'Title':
    #             pass

    #         elif cell.value == 'Values'

    # ontology = {}

    # for row in ws.iter_rows(min_row=2, max_col=12, max_row=45):
    #     object_features = []
    #     for cell in row:
    #         object_features.append(cell.value)

    #     entity = object_features.pop(1)
    #     ontology[entity] = object_features[0:]

    # return ontology


def main():
    ontology_parser()


if __name__ == '__main__':
    main()
