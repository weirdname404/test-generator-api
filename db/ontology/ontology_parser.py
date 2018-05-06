#!/usr/bin/env python
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
# from sqlalchemy import create_engine
# from pprint import pprint

# Instead of DB we use json file to test the module
import json

# engine = create_engine('sqlite:///:memory:', echo=True)


def parse_ontology(file_name):
    # Load in the workbook
    # This version parses only MVContext
    ws = load_workbook(file_name)['MVContext']
    ontology = {}

    for row in ws.iter_rows(min_row=2, max_col=12, max_row=45):
        object_features = []
        for cell in row:
            object_features.append(cell.value)

        entity = object_features.pop(1)
        ontology[entity] = object_features[0:]

    return ontology


def main():
    FILE_NAME = './ontology.xlsm'
    with open('ontology.json', 'w', encoding='utf8') as json_file:
        json.dump(parse_ontology(FILE_NAME), json_file, ensure_ascii=False)


if __name__ == '__main__':
    main()
